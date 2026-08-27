from openpyxl import load_workbook
from openpyxl.chart import BarChart , Reference
from openpyxl.utils  import get_column_letter
from openpyxl.styles import Font
import os
import sys

app_path = os.path.dirname(sys.executable)

#month = 'February'
month = input("Introduce month :")

input_path = os.path.join(app_path,'pivot_table.xlsx')
wb = load_workbook(input_path)
sheet = wb['Report']
sheet.sheet_format.defaultColWidth = 15

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet , min_col=min_column+1 , max_col=max_column , min_row=min_row , max_row=max_row)
categories = Reference(sheet , min_col=min_column , max_col=min_column , min_row=min_row+1 , max_row=max_row)

barchart.add_data(data , titles_from_data = True)
barchart.set_categories(categories)

sheet.add_chart(barchart,"B12")

barchart.title = "Sales by product value"
barchart.varyColors = True
barchart.style = 10

for i in range(min_column+1,max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

sheet[f'{get_column_letter(min_column)}{max_row+1}'] = 'Total'

sheet['A1']='Sales Report'
sheet['A2']= month
sheet['A1'].font = Font('Arial' , bold=True , size=20)
sheet['A2'].font = Font('Arial' , bold=True , size=10)

out_path = os.path.join(app_path,f'Report-{month}.xlsx')

wb.save(out_path)